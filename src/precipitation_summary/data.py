#!/usr/bin/env python3

import collections as cl
import datetime    as dt
import itertools   as it
import functools   as ft

import openpyxl       as xl
import openpyxl.utils as xl_utils

import src.precipitation_summary.rain as rain
import src.precipitation_summary.util as util


LEFT_ALIGN = xl.styles.Alignment(horizontal='left')


def parse_data(rain_row_it):
    result    = cl.defaultdict(list)
    cell_val  = lambda c: c.value or 0
    is_outage = lambda c: c.value == None
    for r in rain_row_it:
        for c in r:
            val = (cell_val(c), is_outage(c))
            result[c.column_letter].append(val)
    return result


def iter_parsed(parsed):
    vals    = util.flatten(parsed[k] for k in sorted(parsed.keys()))
    flatten = lambda i: tuple(util.value_chain(i))
    return map(flatten, enumerate(vals))


def from_workbook(fpath):
    workbook  = xl.load_workbook(filename=fpath)
    worksheet = workbook.active
    data_it   = map(ft.partial(util.drop, 4), util.drop(4, worksheet.iter_rows()))
    return (worksheet['A1'].value, iter_parsed(parse_data(data_it)))


def make_cell(worksheet, cell_style, val):
    cell       = xl.cell.Cell(worksheet)
    cell.value = val
    for k, v in cell_style.items():
        setattr(cell, k, v)
    return cell


def set_column_width(worksheet, labels_it):
    cl_it = map(xl_utils.get_column_letter, it.count(1))
    for c, l in zip(cl_it, labels_it):
        worksheet.column_dimensions[c].width = max(10, 1.23*len(str(l)))


def write_sheet_labels(worksheet, labels_it):
    style = {'font': xl.styles.Font(name='Calibri', bold=True)}
    worksheet.append(tuple(map(ft.partial(make_cell, worksheet, style), labels_it)))
    set_column_width(worksheet, labels_it)


def make_solid_fill(color):
    return xl.styles.PatternFill('solid', fgColor=color)


def write_timeline_sheet(worksheet, station, event_it):
    fields = cl.OrderedDict({
        'id'                 : lambda i, _: i,
        'datum [YYYY-MM-DD]' : lambda _, d: f'{d[2].year:04}-{d[2].month:02}-{d[2].day:02}',
        'rok'                : lambda _, d: d[2].year,
        'měsíc'              : lambda _, d: d[2].month,
        'den'                : lambda _, d: d[2].day,
        'termín [HH:MM]'     : lambda _, d: f'{d[2].hour:02}:{d[2].minute:02}',
        'SRA10M [mm/10min.]' : lambda _, d: d[1],
    })
    write_sheet_labels(worksheet, [station])
    write_sheet_labels(worksheet, fields.keys())
    fill_it = it.cycle((make_solid_fill('d5f5c6'), make_solid_fill('f2c9a3')))
    for i, e in enumerate(event_it, start=1):
        style = {'alignment': LEFT_ALIGN, 'fill': next(fill_it)}
        vals  = lambda it: map(lambda fn: fn(i, it), fields.values())
        cells = lambda it: map(ft.partial(make_cell, worksheet, style), it)
        for row in map(cells, map(vals, e)):
            worksheet.append(tuple(row))


def write_param_sheet(worksheet, station, event_it):
    fields = cl.OrderedDict({
        'id'                 : lambda i, _: i,
        'datum [YYYY-MM-DD]' : lambda _, e: f'{e[0][2].year:04}-{e[0][2].month:02}-{e[0][2].day:02}',
        'rok'                : lambda _, e: e[0][2].year,
        'měsíc'              : lambda _, e: e[0][2].month,
        'den'                : lambda _, e: e[0][2].day,
        'doba trvání [hod]'  : lambda _, e: \
                round((e[-1][2] - e[0][2])/dt.timedelta(hours=1), 2),
        'celkový úhrn [mm]'  : lambda _, e: \
                round(rain.total_amount(e), 2),
        '20 min. max. [mm]'  : lambda _, e: \
                round(rain.total_max_period(util.minutes(20), e), 2),
        '30 min. max. [mm]'  : lambda _, e: \
                round(rain.total_max_period(util.minutes(30), e), 2),
        'kin. energie [MJ]'  : lambda _, e: \
                round(rain.total_kinetic_energy(e), 4),
        'erozivita [R]'      : lambda _, e: \
                round(rain.total_erosivity(util.minutes(30), e), 4),
    })
    write_sheet_labels(worksheet, [station])
    write_sheet_labels(worksheet, fields.keys())
    fill_it = it.cycle((make_solid_fill('d5f5c6'), make_solid_fill('f2c9a3')))
    for i, e in enumerate(event_it, start=1):
        style = {'alignment': LEFT_ALIGN, 'fill': next(fill_it)}
        vals  = map(lambda fn: fn(i, e), fields.values())
        worksheet.append(tuple(map(ft.partial(make_cell, worksheet, style), vals)))


def write_station_workbook(workbook, name, station, event_it):
    event_it             = tuple(event_it)
    timeline_sheet       = workbook.create_sheet()
    timeline_sheet.title = name
    write_timeline_sheet(timeline_sheet, station, event_it)

    param_sheet       = workbook.create_sheet()
    param_sheet.title = f'{name}_params'
    write_param_sheet(param_sheet, station, event_it)


LGT_RAIN_DESC = "celkový úhrn > 1.27 mm"
MID_RAIN_DESC = "20 min. max. > 8.3 mm nebo celkový úhrn > 12.5 mm."
HVY_RAIN_DESC = "20 min. max. > 8.3 mm a celkový úhrn > 12.5 mm."


def to_station_workbook(fpath, station, events_it, mid_events_it, heavy_events_it):
    workbook = xl.Workbook()
    workbook.remove(workbook.active)
    write_station_workbook(workbook, 'light_rain',  f'{station} ({LGT_RAIN_DESC})', events_it)
    write_station_workbook(workbook, 'middle_rain', f'{station} ({MID_RAIN_DESC})', mid_events_it)
    write_station_workbook(workbook, 'heavy_rain',  f'{station} ({HVY_RAIN_DESC})', heavy_events_it)
    workbook.save(fpath)


def make_append_stat_sheet(worksheet, heading):
    fields = cl.OrderedDict({
        'id'                       : lambda i, _, __: i,
        'stanice'                  : lambda _, s, __: s,
        'počet srážek za 10let'    : lambda _, __, e: len(e),
        'suma SRA10M [mm/10let]'   : lambda _, __, e: \
                round(sum(map(rain.total_amount, e)), 2),
        'suma erozivity [R/10let]' : lambda _, __, e: \
                round(sum(map(ft.partial(rain.total_erosivity, util.minutes(30)), e)), 4),
    })
    write_sheet_labels(worksheet, [heading])
    write_sheet_labels(worksheet, fields.keys())
    fill_it = it.cycle((make_solid_fill('d5f5c6'), make_solid_fill('f2c9a3')))
    def append_stat_sheet(idx, station, event_it):
        style = {'alignment': LEFT_ALIGN, 'fill': next(fill_it)}
        vals  = map(lambda fn: fn(idx, station, event_it), fields.values())
        worksheet.append(tuple(map(ft.partial(make_cell, worksheet, style), vals)))
    return append_stat_sheet


def make_append_montly_stat_sheet(worksheet, heading):
    fields = cl.OrderedDict({
        'id'                       : lambda i, _, __: i,
        'stanice'                  : lambda _, s, __: s,
        'měsíc'                    : lambda _, __, m: m[0],
        'počet srážek za měsíc'    : lambda _, __, m: len(m[1]),
        'suma SRA10M [mm/měsíc]'   : lambda _, __, m: \
                round(sum(map(rain.total_amount, m[1])), 2),
        'suma erozivity [R/měsíc]' : lambda _, __, m: \
                round(sum(map(ft.partial(rain.total_erosivity, util.minutes(30)), m[1])), 4),
    })
    write_sheet_labels(worksheet, [heading])
    write_sheet_labels(worksheet, fields.keys())
    fill_it = it.cycle((make_solid_fill('d5f5c6'), make_solid_fill('f2c9a3')))
    def append_montly_stat_sheet(idx, station, monthly_it):
        style = lambda: {'alignment': LEFT_ALIGN, 'fill': next(fill_it)}
        vals  = lambda it: map(lambda fn: fn(idx, station, it), fields.values())
        cells = lambda it: map(ft.partial(make_cell, worksheet, style()), it)
        for row in map(cells, map(vals, monthly_it)):
            worksheet.append(tuple(row))
    return append_montly_stat_sheet


def make_append_stat_workbook(workbook, name, heading):
    stat_sheet        = workbook.create_sheet()
    stat_sheet.title  = name
    append_stat_sheet = make_append_stat_sheet(stat_sheet, heading)

    monthly_stat_sheet       = workbook.create_sheet()
    monthly_stat_sheet.title = f'{name}_monthly'
    append_montly_stat_sheet = make_append_montly_stat_sheet(monthly_stat_sheet, heading)
    def append_stat_workbook(idx, station, event_it):
        event_it = tuple(event_it)
        append_stat_sheet       (idx, station, event_it)
        append_montly_stat_sheet(idx, station, rain.iter_monthly(event_it))
    return append_stat_workbook


def make_to_stat_workbook(fpath):
    workbook          = xl.Workbook()
    workbook.remove(workbook.active)
    append_rain       = make_append_stat_workbook(workbook, 'light_rain',  LGT_RAIN_DESC)
    append_mild_rain  = make_append_stat_workbook(workbook, 'middle_rain', MID_RAIN_DESC)
    append_heavy_rain = make_append_stat_workbook(workbook, 'heavy_rain',  HVY_RAIN_DESC)
    def to_stat_workbook(idx, station, light_events_it, mid_events_it, heavy_event_it):
        append_rain      (idx, station, light_events_it)
        append_mild_rain (idx, station, mid_events_it)
        append_heavy_rain(idx, station, heavy_event_it)
        workbook.save(fpath)
    return to_stat_workbook

