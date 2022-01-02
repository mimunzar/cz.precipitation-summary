#!/usr/bin/env python3

import openpyxl as xl

import src.vahove_srazkomery.rain as rain

import collections as cl
import datetime    as dt
import functools   as ft
import itertools   as it


def iter_rain_rows(rows_it):
    rain_cols = lambda r: it.islice(r, 4, None)
    return map(rain_cols, it.islice(rows_it, 4, None))


def iter_rain_rows_data(rain_row_it):
    col_vals = cl.defaultdict(list)
    for r in rain_row_it:
        for c in r:
            col_vals[c.column_letter].append(c.value or 0)
    joined_cols = (col_vals[k] for k in sorted(col_vals.keys()))
    return enumerate(it.chain.from_iterable(joined_cols))


def from_file(fpath):
    wb = xl.load_workbook(filename=fpath, read_only=True)
    ws = wb.worksheets[0]
    iter_data  = iter_rain_rows_data(iter_rain_rows(ws.iter_rows()))
    return (ws['A1'].value, iter_data)


def make_formatter(fields, fn_time):
    delta  = dt.timedelta(minutes=10)
    offset = int((dt.datetime(2010, 1, 1) - dt.datetime(1970, 1, 1))/delta)
    def formatter(idx, val):
        d = dt.datetime.fromtimestamp((offset + fn_time(val))*delta.seconds)
        return tuple(fn(idx, d, val) for fn in fields.values())
    return formatter


def make_cell(ws, val, style={}):
    c       = xl.cell.Cell(ws)
    c.value = val
    for k, v in style.items():
        setattr(c, k, v)
    return c


def write_sheet_header(ws, station, labels):
    font_setting = xl.styles.Font(name='Calibri', bold=True)
    val_in_bold  = lambda x: make_cell(ws, x, style={'font': font_setting})
    row_in_bold  = lambda i: list(map(val_in_bold, i))
    ws.append(row_in_bold([station]))
    ws.append(row_in_bold(labels))


def write_sheet_data(ws, fn_event_to_rows, event_it):
    left          = xl.styles.Alignment(horizontal='left')
    fill          = lambda c: xl.styles.PatternFill('solid', fgColor=c)
    val_with_fill = lambda f, x: make_cell(ws, x,style={'fill': f, 'alignment': left})
    row_with_fill = lambda f, i: [val_with_fill(f, x) for x in i]
    fill_event_it = zip(it.cycle([fill('d5f5c6'), fill('f2c9a3')]), event_it)
    for i, (fill, event) in enumerate(fill_event_it, start=1):
        for row in fn_event_to_rows(i, event):
            ws.append(row_with_fill(fill, row))


def write_selected_events(ws, station, event_it):
    fields = cl.OrderedDict({
        'id'    : lambda i, _, __: i,
        'rok'   : lambda _, d, __: d.year,
        'měsíc' : lambda _, d, __: d.month,
        'den'   : lambda _, d, __: d.day,
        'termín': lambda _, d, __: f'{d.hour:02}:{d.minute:02}',
        'SRA10M': lambda _, __, r: r[1],
    })
    write_sheet_header(ws, station, fields.keys())
    formatter     = make_formatter(fields, lambda v: v[0])
    event_to_rows = lambda i, e: (formatter(i, v) for v in e)
    write_sheet_data(ws, event_to_rows, event_it)


def minutes(n):
    return n//10


def format_event_duration(n):
    delta = n*dt.timedelta(minutes=10)
    time  = (dt.datetime.min + delta).time()
    return f'{delta.days:02}:{time.hour:02}:{time.minute:02}'


def write_events_sumary(ws, station, event_it):
    fields = cl.OrderedDict({
        'id'                : lambda i, _, __: i,
        'rok'               : lambda _, d, __: d.year,
        'měsíc'             : lambda _, d, __: d.month,
        'den'               : lambda _, d, __: d.day,
        'trvání [DD:HH:MM]' : lambda _, __, e: format_event_duration(len(e)),
        'celkový úhrn [mm]' : lambda _, __, e: round(rain.total_amount(e), 2),
        '20 min. max. [mm]' : lambda _, __, e: \
                round(rain.total_amount(rain.max_period(minutes(20), e)), 2),
        '30 min. max. [mm]' : lambda _, __, e: \
                round(rain.total_amount(rain.max_period(minutes(30), e)), 2),
    })
    write_sheet_header(ws, station, fields.keys())
    formatter     = make_formatter(fields, lambda e: e[0][0])
    event_to_rows = lambda i, e: [formatter(i, e)]
    write_sheet_data(ws, event_to_rows, event_it)


def write_rain_sheet(fpath, station, event_it):
    wb                   = xl.Workbook()
    it_1, it_2           = it.tee(event_it)
    selected_sheet       = wb.active
    selected_sheet.title = 'selected_events'
    write_selected_events(selected_sheet, station, it_1)

    summary_sheet       = wb.create_sheet()
    summary_sheet.title = 'events_summary'
    write_events_sumary(summary_sheet, station, it_2)
    wb.save(fpath)

